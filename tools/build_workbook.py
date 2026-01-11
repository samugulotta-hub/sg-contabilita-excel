from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "contabilita_SG_Impianti_2026.xlsx"


def build_workbook(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contabilita 2026"

    title = "Contabilita SG Impianti - 2026"
    sheet["A1"] = title
    sheet["A1"].font = Font(size=14, bold=True)

    headers = ["Data", "Voce", "Descrizione", "Entrate", "Uscite", "Note"]
    sheet.append(headers)

    sheet.freeze_panes = "A3"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 40
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 30

    sheet.append([date.today().isoformat(), "", "", "", "", ""])
# =========================
# SANITY CHECK OBBLIGATORIO
# =========================
sheetnames = wb.sheetnames
print("Workbook sheets count:", len(sheetnames))
print("Workbook sheets:", sheetnames)

if len(sheetnames) < 10:
    raise Exception(
        f"ERRORE GRAVE: creati solo {len(sheetnames)} fogli. Devono essere almeno 10."
    )

    workbook.save(output_path)


if __name__ == "__main__":
    build_workbook(OUTPUT_FILE)
